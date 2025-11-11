"""
Fire Department Management System - SQLite Version
Clean rewrite using SQLite database instead of JSON files
"""

from flask import Flask, render_template, request, redirect, url_for, session, flash, send_file, jsonify, make_response
from datetime import datetime, timedelta
import pytz
from io import StringIO, BytesIO
import csv
import os
import logging
from logging.handlers import RotatingFileHandler
from dotenv import load_dotenv
import db_helpers
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT

# Load environment variables
load_dotenv(os.path.join(os.path.dirname(__file__), '..', '.env'))

# Initialize Flask application
app = Flask(__name__)
app.secret_key = os.getenv('FLASK_SECRET_KEY')
ADMIN_USERNAME = os.getenv('ADMIN_USERNAME')
ADMIN_PASSWORD = os.getenv('ADMIN_PASSWORD')

# Validate required environment variables
if not all([app.secret_key, ADMIN_USERNAME, ADMIN_PASSWORD]):
    raise ValueError("Missing required environment variables. Check your .env file.")

# Enhanced logging configuration
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        RotatingFileHandler('firefighter.log', maxBytes=1000000, backupCount=5),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# Timezone configuration
central = pytz.timezone('America/Chicago')

# Template filters
@app.template_filter('fromisoformat')
def fromisoformat_filter(date_string):
    return datetime.fromisoformat(date_string)

def format_log_time(log_time):
    """Format log time to a more readable format with correct timezone"""
    dt = datetime.fromisoformat(log_time)
    if dt.tzinfo is None:
        dt = pytz.utc.localize(dt)
    central_time = dt.astimezone(central)
    return central_time.strftime('%Y-%m-%d %H:%M:%S')

app.jinja_env.filters['format_log_time'] = format_log_time

# ========== UTILITY FUNCTIONS ==========

def get_form_value(form, key, default='', required=False):
    """
    Safely get form value with optional default and required validation
    """
    value = form.get(key, default)
    if required and not value:
        raise ValueError(f"Required field '{key}' is missing")
    return value.strip() if isinstance(value, str) else value

# ========== ERROR HANDLERS ==========

@app.errorhandler(404)
def not_found_error(error):
    """Handle 404 errors"""
    flash('Page not found.')
    return redirect(url_for('index'))

@app.errorhandler(500)
def internal_error(error):
    """Handle 500 errors"""
    logger.error(f"Internal server error: {str(error)}")
    flash('An internal error occurred. Please try again later.')
    return redirect(url_for('index'))

# ========== ROUTES ==========

@app.route('/')
def index():
    """Home page - clock in/out interface"""
    # Auto-checkout stale logs (over 12 hours)
    db_helpers.auto_checkout_stale_logs()

    show_new_user_form = session.get('show_new_user_form', False)
    session.pop('show_new_user_form', None)
    categories = [cat['name'] for cat in db_helpers.get_all_categories()]
    return render_template('index.html', show_new_user_form=show_new_user_form, categories=categories)

@app.route('/kiosk')
def kiosk():
    """Kiosk mode - simplified check in/out interface for iPad"""
    # Auto-checkout stale logs (over 12 hours)
    db_helpers.auto_checkout_stale_logs()

    firefighters = db_helpers.get_all_firefighters()
    categories = [cat['name'] for cat in db_helpers.get_all_categories()]

    # Put "Work night" first as default (most common activity)
    # Handle both "Work night" and "Work Night" (case variations)
    work_night = None
    for cat in categories[:]:
        if cat.lower() == 'work night':
            work_night = cat
            categories.remove(cat)
            break

    if work_night:
        categories.insert(0, work_night)

    # Get current status for each firefighter
    for ff in firefighters:
        latest = db_helpers.get_latest_time_log(ff['fireman_number'])
        if latest and latest['clock_out'] is None:
            ff['is_checked_in'] = True
            ff['current_activity'] = latest.get('activity', 'On Duty')
        else:
            ff['is_checked_in'] = False
            ff['current_activity'] = None

    return render_template('kiosk.html', firefighters=firefighters, categories=categories)

@app.route('/register', methods=['POST'])
def register():
    """Register a new firefighter"""
    try:
        full_name = request.form['full_name']
        fireman_number = request.form['fireman_number']

        result = db_helpers.create_firefighter(fireman_number, full_name)

        if result:
            flash(f'Fireman {full_name} registered successfully!')
            logger.info(f"New firefighter registered: {full_name}")
        else:
            flash(f'Fireman {full_name} already exists!')

    except Exception as e:
        logger.error(f"Registration error: {str(e)}")
        flash('An error occurred during registration.')

    return redirect(url_for('index'))

@app.route('/welcome')
def welcome():
    """Welcome screen after check-in with optional inspection QR"""
    firefighter_name = request.args.get('name', 'Firefighter')
    activity = request.args.get('activity', 'your shift')

    # Check if vehicles need inspection
    alerts = db_helpers.get_all_alerts()
    show_inspection_qr = len(alerts.get('inspections_overdue', [])) > 0

    # Get base URL for QR code
    base_url = request.url_root.rstrip('/')

    return render_template('welcome.html',
                         firefighter_name=firefighter_name,
                         activity=activity,
                         show_inspection_qr=show_inspection_qr,
                         base_url=base_url)

@app.route('/clock_in', methods=['POST'])
def clock_in():
    """Clock in a firefighter"""
    try:
        fireman_number = request.form['username']
        activity = request.form['activity']
        other_activity = request.form.get('other_activity', '').strip()

        if activity == "Other" and other_activity:
            activity = other_activity

        # Check if firefighter exists
        firefighter = db_helpers.get_firefighter_by_number(fireman_number)

        if not firefighter:
            flash('You need to register before clocking in!')
            session['show_new_user_form'] = True
            return redirect(url_for('index'))

        # Clock in
        success, message = db_helpers.clock_in(fireman_number, activity)

        if success:
            flash(f'Fireman {firefighter["full_name"]} clocked in for {activity}!')
            logger.info(f"Clock in: {firefighter['full_name']} - {activity}")
        else:
            flash(f'Error: {message}')

    except Exception as e:
        logger.error(f"Clock in error: {str(e)}")
        flash('An error occurred while clocking in.')

    return redirect(url_for('index'))

@app.route('/clock_out', methods=['POST'])
def clock_out():
    """Clock out a firefighter"""
    try:
        fireman_number = request.form['username']

        # Check if firefighter exists
        firefighter = db_helpers.get_firefighter_by_number(fireman_number)

        if not firefighter:
            flash('You need to register before clocking out!')
            return redirect(url_for('index'))

        # Clock out
        success, message = db_helpers.clock_out(fireman_number)

        if success:
            flash(f'Fireman {firefighter["full_name"]} {message}')
            logger.info(f"Clock out: {firefighter['full_name']} - {message}")
        else:
            flash(f'Error: {message}')

    except Exception as e:
        logger.error(f"Clock out error: {str(e)}")
        flash('An error occurred while clocking out.')

    return redirect(url_for('index'))

@app.route('/clock_in_out', methods=['POST'])
def clock_in_out():
    """Unified clock in/out for kiosk mode - returns JSON"""
    try:
        firefighter_number = request.form['firefighter_number']
        action = request.form['action']
        activity = request.form.get('activity', '')

        # Check if firefighter exists
        firefighter = db_helpers.get_firefighter_by_number(firefighter_number)

        if not firefighter:
            return jsonify({
                'success': False,
                'message': 'Firefighter not found!',
                'needs_registration': True,
                'firefighter_number': firefighter_number
            })

        if action == 'checkin':
            success, message = db_helpers.clock_in(firefighter_number, activity)
            if success:
                logger.info(f"Kiosk clock in: {firefighter['full_name']} - {activity}")
                # Redirect to welcome screen instead of returning JSON
                return jsonify({
                    'success': True,
                    'redirect': url_for('welcome', name=firefighter['full_name'], activity=activity)
                })
            else:
                return jsonify({'success': False, 'message': message})

        elif action == 'checkout':
            success, message = db_helpers.clock_out(firefighter_number)
            if success:
                logger.info(f"Kiosk clock out: {firefighter['full_name']}")
                return jsonify({'success': True, 'message': 'Checked out successfully'})
            else:
                return jsonify({'success': False, 'message': message})

        else:
            return jsonify({'success': False, 'message': 'Invalid action'})

    except Exception as e:
        logger.error(f"Kiosk clock in/out error: {str(e)}", exc_info=True)
        return jsonify({'success': False, 'message': 'An error occurred'})

@app.route('/kiosk_register', methods=['POST'])
def kiosk_register():
    """Register a new firefighter from kiosk - returns JSON"""
    try:
        full_name = request.form['full_name']
        fireman_number = request.form['fireman_number']

        # Register the firefighter
        result = db_helpers.create_firefighter(fireman_number, full_name)

        if result:
            logger.info(f"Kiosk registration: {full_name} (#{fireman_number})")
            return jsonify({'success': True, 'message': f'{full_name} registered successfully!'})
        else:
            return jsonify({'success': False, 'message': f'{full_name} already exists!'})

    except Exception as e:
        logger.error(f"Kiosk registration error: {str(e)}", exc_info=True)
        return jsonify({'success': False, 'message': 'An error occurred during registration'})

# ========== ADMIN ROUTES ==========

@app.route('/admin', methods=['GET'])
def admin():
    """Admin login page"""
    if session.get('logged_in'):
        return redirect(url_for('admin_panel'))
    return render_template('admin_login.html')

@app.route('/admin', methods=['POST'])
def admin_auth():
    """Admin authentication"""
    username = request.form['username']
    password = request.form['password']

    if username == ADMIN_USERNAME and password == ADMIN_PASSWORD:
        session['logged_in'] = True
        logger.info("Admin login successful")
        return redirect(url_for('admin_panel'))

    logger.warning(f"Failed admin login attempt with username: {username}")
    flash('Invalid credentials!')
    return redirect(url_for('admin'))

@app.route('/admin_panel')
def admin_panel():
    """Admin panel - manage firefighters and view logs"""
    if not session.get('logged_in'):
        flash('Please log in first!')
        return redirect(url_for('admin'))

    # Get all data
    firefighters_list = db_helpers.get_all_firefighters()
    categories_list = [cat['name'] for cat in db_helpers.get_all_categories()]

    # Convert to format expected by template
    user_data = {}
    all_logs = []  # Flat list of all logs for sorting

    for ff in firefighters_list:
        logs = db_helpers.get_firefighter_logs(ff['fireman_number'])
        user_data[ff['fireman_number']] = {
            'full_name': ff['full_name'],
            'hours': ff['total_hours'] if ff['total_hours'] is not None else 0.0,
            'logs': logs
        }

        # Add logs to flat list with firefighter info attached
        for log in logs:
            all_logs.append({
                'fireman_number': ff['fireman_number'],
                'full_name': ff['full_name'],
                'log': log
            })

    # Sort all logs by time_in descending (most recent first)
    # Use empty string for None values so they sort to the end
    try:
        all_logs.sort(key=lambda x: x['log'].get('time_in', '') or '', reverse=True)
    except Exception as e:
        logger.error(f"Error sorting logs: {str(e)}", exc_info=True)
        # If sorting fails, just use unsorted list

    return render_template('admin.html', user_data=user_data, categories=categories_list, all_logs=all_logs)

@app.route('/update_hours', methods=['POST'])
def update_hours():
    """Manually add hours for a firefighter"""
    if not session.get('logged_in'):
        flash('Please log in first!')
        return redirect(url_for('admin'))

    try:
        fireman_number = request.form['fireman_number']
        activity = request.form['activity']
        log_date = request.form['log_date']
        time_in = request.form['time_in']
        time_out = request.form['time_out']

        success, message = db_helpers.add_manual_hours(
            fireman_number, activity, log_date, time_in, time_out
        )

        if success:
            firefighter = db_helpers.get_firefighter_by_number(fireman_number)
            flash(f'Created a new log for {firefighter["full_name"]}: {message}')
            logger.info(f"Manual hours update: {firefighter['full_name']} - {message}")
        else:
            flash(f'Error: {message}')

    except Exception as e:
        logger.error(f"Update hours error: {str(e)}")
        flash('An error occurred while updating hours.')

    return redirect(url_for('admin_panel'))

@app.route('/edit_firefighter', methods=['POST'])
def edit_firefighter():
    """Edit firefighter information"""
    if not session.get('logged_in'):
        flash('Please log in first!')
        return redirect(url_for('admin'))

    try:
        fireman_number = request.form['fireman_number']
        new_fireman_number = request.form['new_fireman_number']
        new_full_name = request.form['full_name']

        db_helpers.update_firefighter(fireman_number, new_fireman_number, new_full_name)

        flash('Firefighter information updated successfully!')
        logger.info(f"Firefighter edited: {new_full_name}")

    except Exception as e:
        logger.error(f"Edit firefighter error: {str(e)}")
        flash('An error occurred while updating firefighter information.')

    return redirect(url_for('admin_panel'))

@app.route('/delete_firefighter', methods=['POST'])
def delete_firefighter():
    """Delete a firefighter"""
    if not session.get('logged_in'):
        flash('Please log in first!')
        return redirect(url_for('admin'))

    try:
        fireman_number = request.form['fireman_number']

        firefighter = db_helpers.get_firefighter_by_number(fireman_number)

        if firefighter:
            db_helpers.delete_firefighter(fireman_number)
            flash(f'Firefighter {firefighter["full_name"]} has been deleted successfully!')
            logger.info(f"Firefighter deleted: {firefighter['full_name']}")
        else:
            flash('Firefighter not found!')

    except Exception as e:
        logger.error(f"Delete firefighter error: {str(e)}")
        flash('An error occurred while deleting the firefighter.')

    return redirect(url_for('admin_panel'))

@app.route('/add_category', methods=['POST'])
def add_category():
    """Add a new activity category"""
    if not session.get('logged_in'):
        flash('Please log in first!')
        return redirect(url_for('admin'))

    try:
        new_category = request.form['new_category']

        result = db_helpers.create_category(new_category)

        if result:
            flash(f'Category "{new_category}" added successfully!')
            logger.info(f"New category added: {new_category}")
        else:
            flash('Invalid or duplicate category!')

    except Exception as e:
        logger.error(f"Add category error: {str(e)}")
        flash('An error occurred while adding the category.')

    return redirect(url_for('admin_panel'))

@app.route('/remove_category', methods=['POST'])
def remove_category():
    """Remove an activity category"""
    if not session.get('logged_in'):
        flash('Please log in first!')
        return redirect(url_for('admin'))

    try:
        category_to_remove = request.form['category_to_remove']

        db_helpers.delete_category(category_to_remove)

        flash(f'Category "{category_to_remove}" removed successfully!')
        logger.info(f"Category removed: {category_to_remove}")

    except Exception as e:
        logger.error(f"Remove category error: {str(e)}")
        flash('An error occurred while removing the category.')

    return redirect(url_for('admin_panel'))

@app.route('/delete_log', methods=['POST'])
def delete_log():
    """Delete a specific log entry"""
    if not session.get('logged_in'):
        flash('Please log in first!')
        return redirect(url_for('admin'))

    try:
        log_id = int(request.form['log_id'])

        success, message = db_helpers.delete_log_by_id(log_id)

        if success:
            flash('Log entry deleted successfully!')
            logger.info(f"Log ID {log_id} deleted")
        else:
            flash(f'Error: {message}')

    except Exception as e:
        logger.error(f"Delete log error: {str(e)}", exc_info=True)
        flash('An error occurred while deleting the log.')

    return redirect(url_for('admin_panel'))

@app.route('/clear_all_logs', methods=['POST'])
def clear_all_logs():
    """Clear all logs and reset hours"""
    if not session.get('logged_in'):
        flash('Please log in first!')
        return redirect(url_for('admin'))

    try:
        db_helpers.clear_all_logs()
        flash('All logs and hours have been cleared successfully!')
        logger.info("All logs cleared")

    except Exception as e:
        logger.error(f"Clear logs error: {str(e)}")
        flash('An error occurred while clearing logs.')

    return redirect(url_for('admin_panel'))

@app.route('/export_data')
def export_data():
    """Export firefighter time logs to CSV (formatted for payroll)"""
    if not session.get('logged_in'):
        flash('Please log in first!')
        return redirect(url_for('admin'))

    try:
        # Get month and year from query parameters
        month = request.args.get('month')
        year = request.args.get('year')

        output = StringIO()
        cw = csv.writer(output)

        # Header row
        if month and year:
            month_name = datetime(int(year), int(month), 1).strftime('%B %Y')
            cw.writerow([f'PAYROLL REPORT - {month_name}'])
        else:
            cw.writerow(['PAYROLL REPORT - ALL TIME'])
        cw.writerow([])  # Blank row

        firefighters = db_helpers.get_all_firefighters()

        for ff in firefighters:
            logs = db_helpers.get_firefighter_logs(ff['fireman_number'])

            # Filter logs by month/year if specified
            filtered_logs = []
            month_total = 0

            for log in logs:
                if log['time_in']:
                    time_in = datetime.fromisoformat(log['time_in'])

                    # Filter by month and year if provided
                    if month and year:
                        if time_in.month != int(month) or time_in.year != int(year):
                            continue

                    filtered_logs.append(log)
                    month_total += log.get('hours_worked', 0) or 0

            # Only show firefighter if they have logs in this period
            if filtered_logs:
                # Firefighter header
                cw.writerow([f"FIREFIGHTER #{ff['fireman_number']} - {ff['full_name']}"])
                cw.writerow(['Date', 'Time In', 'Time Out', 'Activity', 'Hours Worked'])

                # Individual log entries
                for log in filtered_logs:
                    time_in_dt = datetime.fromisoformat(log['time_in']) if log['time_in'] else None
                    time_out_dt = datetime.fromisoformat(log['time_out']) if log['time_out'] else None

                    date_str = time_in_dt.strftime('%Y-%m-%d') if time_in_dt else ''
                    time_in_str = time_in_dt.strftime('%I:%M %p') if time_in_dt else ''
                    time_out_str = time_out_dt.strftime('%I:%M %p') if time_out_dt else 'Still clocked in'
                    hours = f"{log.get('hours_worked', 0):.2f}" if log.get('hours_worked') else '0.00'

                    cw.writerow([
                        date_str,
                        time_in_str,
                        time_out_str,
                        log['type'],
                        hours
                    ])

                # Subtotal for this firefighter
                cw.writerow(['', '', '', 'TOTAL HOURS:', f"{month_total:.2f}"])
                cw.writerow([])  # Blank row between firefighters

        output.seek(0)

        # Generate filename
        if month and year:
            filename = f'payroll_{year}_{int(month):02d}.csv'
        else:
            filename = 'payroll_all_time.csv'

        return send_file(
            BytesIO(output.getvalue().encode()),
            mimetype='text/csv',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        logger.error(f"Export error: {str(e)}")
        flash('An error occurred during export.')
        return redirect(url_for('admin_panel'))

@app.route('/export_inspections')
def export_inspections():
    """Export vehicle inspection logs to CSV"""
    if not session.get('logged_in'):
        flash('Please log in first!')
        return redirect(url_for('admin'))

    try:
        # Get query parameters
        month = request.args.get('month')
        year = request.args.get('year')
        week = request.args.get('week')  # 'current' for current week

        output = StringIO()
        cw = csv.writer(output)

        # Determine date range for filtering
        start_date = None
        end_date = None
        title_suffix = 'ALL TIME'
        filename_suffix = 'all_time'

        if week == 'current':
            # Get current week (Sunday to Saturday)
            now = datetime.now(central)
            # Calculate days since Sunday (0=Monday, 6=Sunday in weekday())
            days_since_sunday = (now.weekday() + 1) % 7
            start_date = (now - timedelta(days=days_since_sunday)).replace(hour=0, minute=0, second=0, microsecond=0)
            end_date = start_date + timedelta(days=6, hours=23, minutes=59, seconds=59)
            title_suffix = f'WEEK OF {start_date.strftime("%B %d, %Y")}'
            filename_suffix = f'week_{start_date.strftime("%Y_%m_%d")}'
        elif month and year:
            title_suffix = datetime(int(year), int(month), 1).strftime('%B %Y')
            filename_suffix = f'{year}_{int(month):02d}'

        # Header row
        cw.writerow([f'VEHICLE INSPECTION REPORT - {title_suffix}'])
        cw.writerow([])
        cw.writerow(['Vehicle Code', 'Vehicle Name', 'Inspection Date', 'Inspector', 'Result', 'Notes'])

        vehicles = db_helpers.get_all_vehicles()

        for vehicle in vehicles:
            history = db_helpers.get_vehicle_inspection_history(vehicle['id'], limit=1000)

            for inspection in history:
                inspection_date = datetime.fromisoformat(inspection['date'])

                # Filter by date range
                if week == 'current':
                    if not (start_date <= inspection_date <= end_date):
                        continue
                elif month and year:
                    if inspection_date.month != int(month) or inspection_date.year != int(year):
                        continue

                date_str = inspection_date.strftime('%Y-%m-%d %I:%M %p')
                result = 'PASSED' if inspection['passed'] else 'FAILED'

                cw.writerow([
                    vehicle['code'],
                    vehicle['name'],
                    date_str,
                    inspection['inspector'],
                    result,
                    inspection.get('notes', '')
                ])

        output.seek(0)

        # Generate filename
        filename = f'inspections_{filename_suffix}.csv'

        return send_file(
            BytesIO(output.getvalue().encode()),
            mimetype='text/csv',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        logger.error(f"Inspection export error: {str(e)}")
        flash('An error occurred during inspection export.')
        return redirect(url_for('admin_panel'))

@app.route('/export_data_pdf')
def export_data_pdf():
    """Export firefighter time logs to PDF (formatted for payroll)"""
    if not session.get('logged_in'):
        flash('Please log in first!')
        return redirect(url_for('admin'))

    try:
        month = request.args.get('month')
        year = request.args.get('year')

        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(letter), topMargin=0.5*inch, bottomMargin=0.5*inch)
        elements = []
        styles = getSampleStyleSheet()

        # Title
        title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], alignment=TA_CENTER, fontSize=16, spaceAfter=20)
        if month and year:
            month_name = datetime(int(year), int(month), 1).strftime('%B %Y')
            title = Paragraph(f"<b>PAYROLL REPORT - {month_name}</b>", title_style)
        else:
            title = Paragraph("<b>PAYROLL REPORT - ALL TIME</b>", title_style)
        elements.append(title)
        elements.append(Spacer(1, 0.25*inch))

        firefighters = db_helpers.get_all_firefighters()

        for ff in firefighters:
            logs = db_helpers.get_firefighter_logs(ff['fireman_number'])
            filtered_logs = []
            month_total = 0

            for log in logs:
                if log['time_in']:
                    time_in = datetime.fromisoformat(log['time_in'])
                    if month and year:
                        if time_in.month != int(month) or time_in.year != int(year):
                            continue
                    filtered_logs.append(log)
                    month_total += log.get('hours_worked', 0) or 0

            if filtered_logs:
                # Firefighter header
                ff_header = Paragraph(f"<b>FIREFIGHTER #{ff['fireman_number']} - {ff['full_name']}</b>", styles['Heading2'])
                elements.append(ff_header)
                elements.append(Spacer(1, 0.1*inch))

                # Table data
                data = [['Date', 'Time In', 'Time Out', 'Activity', 'Hours Worked']]
                for log in filtered_logs:
                    time_in_dt = datetime.fromisoformat(log['time_in']) if log['time_in'] else None
                    time_out_dt = datetime.fromisoformat(log['time_out']) if log['time_out'] else None
                    date_str = time_in_dt.strftime('%Y-%m-%d') if time_in_dt else ''
                    time_in_str = time_in_dt.strftime('%I:%M %p') if time_in_dt else ''
                    time_out_str = time_out_dt.strftime('%I:%M %p') if time_out_dt else 'Still clocked in'
                    hours = f"{log.get('hours_worked', 0):.2f}" if log.get('hours_worked') else '0.00'
                    data.append([date_str, time_in_str, time_out_str, log['type'], hours])

                # Total row
                data.append(['', '', '', 'TOTAL HOURS:', f"{month_total:.2f}"])

                table = Table(data, colWidths=[1.5*inch, 1.2*inch, 1.2*inch, 1.5*inch, 1.2*inch])
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 10),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                    ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
                    ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))
                elements.append(table)
                elements.append(Spacer(1, 0.3*inch))

        doc.build(elements)
        buffer.seek(0)

        if month and year:
            filename = f'payroll_{year}_{int(month):02d}.pdf'
        else:
            filename = 'payroll_all_time.pdf'

        return send_file(buffer, mimetype='application/pdf', as_attachment=True, download_name=filename)

    except Exception as e:
        logger.error(f"PDF Export error: {str(e)}")
        flash('An error occurred during PDF export.')
        return redirect(url_for('admin_panel'))

@app.route('/export_inspections_pdf')
def export_inspections_pdf():
    """Export vehicle inspection logs to PDF"""
    if not session.get('logged_in'):
        flash('Please log in first!')
        return redirect(url_for('admin'))

    try:
        month = request.args.get('month')
        year = request.args.get('year')
        week = request.args.get('week')  # 'current' for current week

        # Determine date range for filtering
        start_date = None
        end_date = None
        title_suffix = 'ALL TIME'
        filename_suffix = 'all_time'

        if week == 'current':
            # Get current week (Sunday to Saturday)
            now = datetime.now(central)
            days_since_sunday = (now.weekday() + 1) % 7
            start_date = (now - timedelta(days=days_since_sunday)).replace(hour=0, minute=0, second=0, microsecond=0)
            end_date = start_date + timedelta(days=6, hours=23, minutes=59, seconds=59)
            title_suffix = f'WEEK OF {start_date.strftime("%B %d, %Y")}'
            filename_suffix = f'week_{start_date.strftime("%Y_%m_%d")}'
        elif month and year:
            title_suffix = datetime(int(year), int(month), 1).strftime('%B %Y')
            filename_suffix = f'{year}_{int(month):02d}'

        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(letter), topMargin=0.5*inch, bottomMargin=0.5*inch)
        elements = []
        styles = getSampleStyleSheet()

        # Title
        title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], alignment=TA_CENTER, fontSize=16, spaceAfter=20)
        title = Paragraph(f"<b>VEHICLE INSPECTION REPORT - {title_suffix}</b>", title_style)
        elements.append(title)
        elements.append(Spacer(1, 0.25*inch))

        data = [['Vehicle Code', 'Vehicle Name', 'Inspection Date', 'Inspector', 'Result', 'Notes']]

        vehicles = db_helpers.get_all_vehicles()
        for vehicle in vehicles:
            history = db_helpers.get_vehicle_inspection_history(vehicle['id'], limit=1000)
            for insp in history:
                insp_date = datetime.fromisoformat(insp['date'])

                # Filter by date range
                if week == 'current':
                    if not (start_date <= insp_date <= end_date):
                        continue
                elif month and year:
                    if insp_date.month != int(month) or insp_date.year != int(year):
                        continue

                date_str = insp_date.strftime('%Y-%m-%d %I:%M %p')
                result = 'PASSED' if insp['passed'] else 'FAILED'
                notes = insp.get('notes', '')[:40] + '...' if len(insp.get('notes', '')) > 40 else insp.get('notes', '')
                data.append([
                    vehicle['code'],
                    vehicle['name'],
                    date_str,
                    insp.get('inspector', 'N/A'),
                    result,
                    notes
                ])

        if len(data) > 1:
            table = Table(data, colWidths=[1.2*inch, 1.5*inch, 2*inch, 1.5*inch, 0.8*inch, 2.5*inch])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            elements.append(table)
        else:
            elements.append(Paragraph("No inspections found for this period.", styles['Normal']))

        doc.build(elements)
        buffer.seek(0)

        filename = f'inspections_{filename_suffix}.pdf'

        return send_file(buffer, mimetype='application/pdf', as_attachment=True, download_name=filename)

    except Exception as e:
        logger.error(f"Inspection PDF export error: {str(e)}")
        flash('An error occurred during PDF export.')
        return redirect(url_for('admin_panel'))

@app.route('/export_checklist_items')
def export_checklist_items():
    """Export checklist items to CSV"""
    if not session.get('logged_in'):
        flash('Please log in first!')
        return redirect(url_for('admin'))

    try:
        output = StringIO()
        cw = csv.writer(output)

        cw.writerow(['INSPECTION CHECKLIST ITEMS'])
        cw.writerow([])
        cw.writerow(['ID', 'Description', 'Category', 'Display Order', 'Active'])

        items = db_helpers.get_all_checklist_items()
        for item in items:
            cw.writerow([
                item['id'],
                item['description'],
                item['category'] or 'General',
                item['display_order'],
                'Yes' if item['is_active'] else 'No'
            ])

        output.seek(0)
        filename = f'checklist_items_{datetime.now().strftime("%Y%m%d")}.csv'

        return send_file(
            BytesIO(output.getvalue().encode()),
            mimetype='text/csv',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        logger.error(f"Checklist export error: {str(e)}")
        flash('An error occurred during export.')
        return redirect(url_for('manage_checklist_items'))

@app.route('/export_checklist_items_pdf')
def export_checklist_items_pdf():
    """Export checklist items to PDF"""
    if not session.get('logged_in'):
        flash('Please log in first!')
        return redirect(url_for('admin'))

    try:
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=0.5*inch, bottomMargin=0.5*inch)
        elements = []
        styles = getSampleStyleSheet()

        title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], alignment=TA_CENTER, fontSize=16, spaceAfter=20)
        title = Paragraph("<b>INSPECTION CHECKLIST ITEMS</b>", title_style)
        elements.append(title)
        elements.append(Spacer(1, 0.25*inch))

        data = [['ID', 'Description', 'Category', 'Order', 'Active']]

        items = db_helpers.get_all_checklist_items()
        for item in items:
            data.append([
                str(item['id']),
                item['description'][:60],
                item['category'] or 'General',
                str(item['display_order']),
                'Yes' if item['is_active'] else 'No'
            ])

        table = Table(data, colWidths=[0.5*inch, 4*inch, 1.2*inch, 0.7*inch, 0.7*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (0, 0), (0, -1), 'CENTER'),
            ('ALIGN', (3, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(table)

        doc.build(elements)
        buffer.seek(0)

        filename = f'checklist_items_{datetime.now().strftime("%Y%m%d")}.pdf'
        return send_file(buffer, mimetype='application/pdf', as_attachment=True, download_name=filename)

    except Exception as e:
        logger.error(f"Checklist PDF export error: {str(e)}")
        flash('An error occurred during PDF export.')
        return redirect(url_for('manage_checklist_items'))

@app.route('/export_vehicles')
def export_vehicles():
    """Export vehicles to CSV"""
    if not session.get('logged_in'):
        flash('Please log in first!')
        return redirect(url_for('admin'))

    try:
        output = StringIO()
        cw = csv.writer(output)

        cw.writerow(['VEHICLE FLEET INVENTORY'])
        cw.writerow([])
        cw.writerow(['Code', 'Name', 'Type', 'Year', 'Make', 'Model', 'VIN', 'License Plate', 'Status', 'Notes'])

        vehicles = db_helpers.get_all_vehicles()
        for vehicle in vehicles:
            cw.writerow([
                vehicle['vehicle_code'],
                vehicle['name'],
                vehicle['vehicle_type'] or '',
                vehicle.get('year', ''),
                vehicle.get('make', ''),
                vehicle.get('model', ''),
                vehicle.get('vin', ''),
                vehicle.get('license_plate', ''),
                vehicle['status'],
                vehicle.get('notes', '')
            ])

        output.seek(0)
        filename = f'vehicles_{datetime.now().strftime("%Y%m%d")}.csv'

        return send_file(
            BytesIO(output.getvalue().encode()),
            mimetype='text/csv',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        logger.error(f"Vehicle CSV export error: {str(e)}", exc_info=True)
        flash(f'An error occurred during CSV export: {str(e)}')
        return redirect(url_for('manage_vehicles'))

@app.route('/export_vehicles_pdf')
def export_vehicles_pdf():
    """Export vehicles to PDF"""
    if not session.get('logged_in'):
        flash('Please log in first!')
        return redirect(url_for('admin'))

    try:
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(letter), topMargin=0.5*inch, bottomMargin=0.5*inch)
        elements = []
        styles = getSampleStyleSheet()

        title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], alignment=TA_CENTER, fontSize=16, spaceAfter=20)
        title = Paragraph("<b>VEHICLE FLEET INVENTORY</b>", title_style)
        elements.append(title)
        elements.append(Spacer(1, 0.25*inch))

        data = [['Code', 'Name', 'Type', 'Year', 'Make', 'Model', 'VIN', 'Plate', 'Status']]

        vehicles = db_helpers.get_all_vehicles()
        for vehicle in vehicles:
            data.append([
                vehicle['vehicle_code'],
                vehicle['name'],
                vehicle['vehicle_type'] or '',
                str(vehicle.get('year', '')),
                vehicle.get('make', ''),
                vehicle.get('model', ''),
                vehicle.get('vin', '')[:10] + '...' if vehicle.get('vin') else '',
                vehicle.get('license_plate', ''),
                vehicle['status']
            ])

        table = Table(data, colWidths=[0.7*inch, 1.2*inch, 1*inch, 0.7*inch, 1*inch, 1*inch, 1.3*inch, 1*inch, 0.8*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(table)

        doc.build(elements)
        buffer.seek(0)

        filename = f'vehicles_{datetime.now().strftime("%Y%m%d")}.pdf'
        return send_file(buffer, mimetype='application/pdf', as_attachment=True, download_name=filename)

    except Exception as e:
        logger.error(f"Vehicle PDF export error: {str(e)}", exc_info=True)
        flash(f'An error occurred during PDF export: {str(e)}')
        return redirect(url_for('manage_vehicles'))

@app.route('/export_station_inventory/<int:station_id>')
def export_station_inventory(station_id):
    """Export station inventory to CSV"""
    if not session.get('logged_in'):
        flash('Please log in first!')
        return redirect(url_for('admin'))

    try:
        station = db_helpers.get_station_by_id(station_id) if hasattr(db_helpers, 'get_station_by_id') else None
        if not station:
            stations = db_helpers.get_all_stations()
            station = next((s for s in stations if s['id'] == station_id), None)

        if not station:
            flash('Station not found!')
            return redirect(url_for('inventory_menu'))

        output = StringIO()
        cw = csv.writer(output)

        cw.writerow([f'COMPLETE STATION INVENTORY - {station["name"]}'])
        cw.writerow(['Includes station building and all vehicles'])
        cw.writerow([])

        grand_total = 0

        # Station building inventory
        cw.writerow(['=== STATION BUILDING INVENTORY ==='])
        cw.writerow(['Item Name', 'Category', 'Quantity', 'Unit', 'Cost Per Unit', 'Total Value', 'Location', 'Last Checked'])

        inventory = db_helpers.get_station_inventory(station_id)
        station_total = 0
        for item in inventory:
            cost_per_unit = item.get('cost_per_unit', 0) or 0
            quantity = item.get('quantity', 0) or 0
            item_value = cost_per_unit * quantity
            station_total += item_value

            cw.writerow([
                item.get('name', ''),
                item.get('category', ''),
                quantity,
                item.get('unit_of_measure', ''),
                f"${cost_per_unit:.2f}",
                f"${item_value:.2f}",
                station['name'],
                item.get('last_checked', '')
            ])

        cw.writerow(['', '', '', '', 'STATION SUBTOTAL:', f"${station_total:.2f}", '', ''])
        cw.writerow([])
        grand_total += station_total

        # Vehicle inventories
        vehicles = db_helpers.get_vehicles_by_station(station_id)
        for vehicle in vehicles:
            cw.writerow([f'=== {vehicle["name"]} ({vehicle["vehicle_code"]}) ==='])
            cw.writerow(['Item Name', 'Category', 'Quantity', 'Unit', 'Cost Per Unit', 'Total Value', 'Location', 'Last Checked'])

            vehicle_inventory = db_helpers.get_vehicle_inventory(vehicle['id'])
            vehicle_total = 0
            for item in vehicle_inventory:
                cost_per_unit = item.get('cost_per_unit', 0) or 0
                quantity = item.get('quantity', 0) or 0
                item_value = cost_per_unit * quantity
                vehicle_total += item_value

                cw.writerow([
                    item.get('name', ''),
                    item.get('category', ''),
                    quantity,
                    item.get('unit_of_measure', ''),
                    f"${cost_per_unit:.2f}",
                    f"${item_value:.2f}",
                    f'{vehicle["name"]} (Truck)',
                    item.get('last_checked', '')
                ])

            cw.writerow(['', '', '', '', f'{vehicle["name"]} SUBTOTAL:', f"${vehicle_total:.2f}", '', ''])
            cw.writerow([])
            grand_total += vehicle_total

        # Grand total
        cw.writerow(['', '', '', '', 'GRAND TOTAL (STATION + ALL VEHICLES):', f"${grand_total:.2f}", '', ''])

        output.seek(0)
        filename = f'station_{station_id}_inventory_{datetime.now().strftime("%Y%m%d")}.csv'

        return send_file(
            BytesIO(output.getvalue().encode()),
            mimetype='text/csv',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        logger.error(f"Station inventory CSV export error: {str(e)}", exc_info=True)
        flash(f'An error occurred during CSV export: {str(e)}')
        return redirect(url_for('station_inventory', station_id=station_id))

@app.route('/export_station_inventory_pdf/<int:station_id>')
def export_station_inventory_pdf(station_id):
    """Export station inventory to PDF"""
    if not session.get('logged_in'):
        flash('Please log in first!')
        return redirect(url_for('admin'))

    try:
        station = db_helpers.get_station_by_id(station_id) if hasattr(db_helpers, 'get_station_by_id') else None
        if not station:
            stations = db_helpers.get_all_stations()
            station = next((s for s in stations if s['id'] == station_id), None)

        if not station:
            flash('Station not found!')
            return redirect(url_for('inventory_menu'))

        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(letter), topMargin=0.5*inch, bottomMargin=0.5*inch)
        elements = []
        styles = getSampleStyleSheet()

        title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], alignment=TA_CENTER, fontSize=16, spaceAfter=20)
        title = Paragraph(f"<b>COMPLETE STATION INVENTORY - {station['name']}</b>", title_style)
        subtitle_style = ParagraphStyle('Subtitle', parent=styles['Normal'], alignment=TA_CENTER, fontSize=12, spaceAfter=20)
        subtitle = Paragraph("<i>Includes station building and all vehicles</i>", subtitle_style)
        elements.append(title)
        elements.append(subtitle)
        elements.append(Spacer(1, 0.25*inch))

        grand_total = 0

        # Station building inventory
        station_header = Paragraph("<b>STATION BUILDING INVENTORY</b>", styles['Heading2'])
        elements.append(station_header)
        elements.append(Spacer(1, 0.1*inch))

        data = [['Item Name', 'Category', 'Qty', 'Unit', 'Cost', 'Value']]
        inventory = db_helpers.get_station_inventory(station_id)
        station_total = 0
        for item in inventory:
            cost_per_unit = item.get('cost_per_unit', 0) or 0
            quantity = item.get('quantity', 0) or 0
            item_value = cost_per_unit * quantity
            station_total += item_value

            data.append([
                item.get('name', '')[:30],
                item.get('category', '')[:15],
                str(quantity),
                item.get('unit_of_measure', '')[:8],
                f"${cost_per_unit:.2f}",
                f"${item_value:.2f}"
            ])

        if len(data) > 1:
            data.append(['', '', '', '', 'SUBTOTAL:', f"${station_total:.2f}"])
            table = Table(data, colWidths=[2.5*inch, 1.5*inch, 0.7*inch, 0.8*inch, 1*inch, 1.2*inch])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
                ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            elements.append(table)
            grand_total += station_total

        elements.append(Spacer(1, 0.3*inch))

        # Vehicle inventories
        vehicles = db_helpers.get_vehicles_by_station(station_id)
        for vehicle in vehicles:
            vehicle_header = Paragraph(f"<b>{vehicle['name']} ({vehicle['vehicle_code']})</b>", styles['Heading2'])
            elements.append(vehicle_header)
            elements.append(Spacer(1, 0.1*inch))

            data = [['Item Name', 'Category', 'Qty', 'Unit', 'Cost', 'Value']]
            vehicle_inventory = db_helpers.get_vehicle_inventory(vehicle['id'])
            vehicle_total = 0
            for item in vehicle_inventory:
                cost_per_unit = item.get('cost_per_unit', 0) or 0
                quantity = item.get('quantity', 0) or 0
                item_value = cost_per_unit * quantity
                vehicle_total += item_value

                data.append([
                    item.get('name', '')[:30],
                    item.get('category', '')[:15],
                    str(quantity),
                    item.get('unit_of_measure', '')[:8],
                    f"${cost_per_unit:.2f}",
                    f"${item_value:.2f}"
                ])

            if len(data) > 1:
                data.append(['', '', '', '', 'SUBTOTAL:', f"${vehicle_total:.2f}"])
                table = Table(data, colWidths=[2.5*inch, 1.5*inch, 0.7*inch, 0.8*inch, 1*inch, 1.2*inch])
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 10),
                    ('FONTSIZE', (0, 1), (-1, -1), 9),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                    ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
                    ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))
                elements.append(table)
                grand_total += vehicle_total
            else:
                elements.append(Paragraph("<i>No inventory items on this vehicle</i>", styles['Normal']))

            elements.append(Spacer(1, 0.3*inch))

        # Grand total
        grand_total_para = Paragraph(f"<b>GRAND TOTAL (STATION + ALL VEHICLES): ${grand_total:.2f}</b>", styles['Heading2'])
        elements.append(grand_total_para)

        doc.build(elements)
        buffer.seek(0)

        filename = f'station_{station_id}_inventory_{datetime.now().strftime("%Y%m%d")}.pdf'
        return send_file(buffer, mimetype='application/pdf', as_attachment=True, download_name=filename)

    except Exception as e:
        logger.error(f"Station inventory PDF export error: {str(e)}", exc_info=True)
        flash(f'An error occurred during PDF export: {str(e)}')
        return redirect(url_for('station_inventory', station_id=station_id))

@app.route('/export_vehicle_inventory/<int:vehicle_id>')
def export_vehicle_inventory(vehicle_id):
    """Export vehicle inventory to CSV"""
    if not session.get('logged_in'):
        flash('Please log in first!')
        return redirect(url_for('admin'))

    try:
        vehicle = db_helpers.get_vehicle_by_id(vehicle_id)
        if not vehicle:
            flash('Vehicle not found!')
            return redirect(url_for('inventory_menu'))

        output = StringIO()
        cw = csv.writer(output)

        cw.writerow([f'VEHICLE INVENTORY - {vehicle["name"]} ({vehicle["code"]})'])
        cw.writerow([])
        cw.writerow(['Item Name', 'Category', 'Quantity', 'Unit', 'Cost Per Unit', 'Total Value', 'Last Checked'])

        inventory = db_helpers.get_vehicle_inventory(vehicle_id)
        total_value = 0
        for item in inventory:
            cost_per_unit = item.get('cost_per_unit', 0) or 0
            quantity = item.get('quantity', 0) or 0
            item_value = cost_per_unit * quantity
            total_value += item_value

            cw.writerow([
                item.get('name', ''),
                item.get('category', ''),
                quantity,
                item.get('unit_of_measure', ''),
                f"${cost_per_unit:.2f}",
                f"${item_value:.2f}",
                item.get('last_checked', '')
            ])

        # Add total row
        cw.writerow([])
        cw.writerow(['', '', '', '', 'TOTAL VALUE:', f"${total_value:.2f}", ''])

        output.seek(0)
        filename = f'vehicle_{vehicle_id}_inventory_{datetime.now().strftime("%Y%m%d")}.csv'

        return send_file(
            BytesIO(output.getvalue().encode()),
            mimetype='text/csv',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        logger.error(f"Vehicle inventory CSV export error: {str(e)}", exc_info=True)
        flash(f'An error occurred during CSV export: {str(e)}')
        return redirect(url_for('vehicle_inventory', vehicle_id=vehicle_id))

@app.route('/export_vehicle_inventory_pdf/<int:vehicle_id>')
def export_vehicle_inventory_pdf(vehicle_id):
    """Export vehicle inventory to PDF"""
    if not session.get('logged_in'):
        flash('Please log in first!')
        return redirect(url_for('admin'))

    try:
        vehicle = db_helpers.get_vehicle_by_id(vehicle_id)
        if not vehicle:
            flash('Vehicle not found!')
            return redirect(url_for('inventory_menu'))

        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(letter), topMargin=0.5*inch, bottomMargin=0.5*inch)
        elements = []
        styles = getSampleStyleSheet()

        title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], alignment=TA_CENTER, fontSize=16, spaceAfter=20)
        title = Paragraph(f"<b>VEHICLE INVENTORY - {vehicle['name']} ({vehicle['code']})</b>", title_style)
        elements.append(title)
        elements.append(Spacer(1, 0.25*inch))

        data = [['Item Name', 'Category', 'Qty', 'Unit', 'Cost', 'Value']]

        inventory = db_helpers.get_vehicle_inventory(vehicle_id)
        total_value = 0
        for item in inventory:
            cost_per_unit = item.get('cost_per_unit', 0) or 0
            quantity = item.get('quantity', 0) or 0
            item_value = cost_per_unit * quantity
            total_value += item_value

            data.append([
                item.get('name', '')[:30],
                item.get('category', '')[:15],
                str(quantity),
                item.get('unit_of_measure', '')[:8],
                f"${cost_per_unit:.2f}",
                f"${item_value:.2f}"
            ])

        # Add total row
        if len(data) > 1:
            data.append(['', '', '', '', 'TOTAL:', f"${total_value:.2f}"])

        if len(data) > 1:
            table = Table(data, colWidths=[2.5*inch, 1.5*inch, 0.7*inch, 0.8*inch, 1*inch, 1.2*inch])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
                ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            elements.append(table)
        else:
            elements.append(Paragraph("No inventory items found for this vehicle.", styles['Normal']))

        doc.build(elements)
        buffer.seek(0)

        filename = f'vehicle_{vehicle_id}_inventory_{datetime.now().strftime("%Y%m%d")}.pdf'
        return send_file(buffer, mimetype='application/pdf', as_attachment=True, download_name=filename)

    except Exception as e:
        logger.error(f"Vehicle inventory PDF export error: {str(e)}", exc_info=True)
        flash(f'An error occurred during PDF export: {str(e)}')
        return redirect(url_for('vehicle_inventory', vehicle_id=vehicle_id))

@app.route('/logout')
def logout():
    """Logout admin"""
    session.pop('logged_in', None)
    flash('Logged out successfully!')
    logger.info("Admin logged out")
    return redirect(url_for('index'))

# ========== DISPLAY ROUTES ==========

@app.route('/display')
def display():
    """Display dashboard - active firefighters, alerts, and recent activity"""
    try:
        # Get station parameter (optional)
        station_id = request.args.get('station', type=int)

        # Auto-checkout stale logs (over 12 hours)
        db_helpers.auto_checkout_stale_logs()

        active_firefighters = db_helpers.get_active_firefighters()
        leaderboard = db_helpers.get_leaderboard()
        vehicles_needing_inspection = db_helpers.get_vehicles_needing_inspection(station_id=station_id)
        alerts = db_helpers.get_all_alerts(station_id=station_id)
        recent_activity = db_helpers.get_recent_activity(limit=15)

        # Get station info if filtering by station
        station = None
        stations = db_helpers.get_all_stations()
        if station_id:
            station = db_helpers.get_station_by_id(station_id)

        # Get base URL for QR codes
        base_url = request.url_root.rstrip('/')

        # Get display settings from database (server-side, not localStorage)
        display_settings = db_helpers.get_all_display_settings()

        logger.info(f"Display page loaded successfully (station_id={station_id})")
        response = make_response(render_template('display.html',
                             active_firefighters=active_firefighters,
                             leaderboard=leaderboard,
                             vehicles_needing_inspection=vehicles_needing_inspection,
                             alerts=alerts,
                             recent_activity=recent_activity,
                             base_url=base_url,
                             display_settings=display_settings,
                             station=station,
                             stations=stations,
                             current_station_id=station_id))

        # Allow this page to be embedded in iframes (for SignPresenter and trusted origins)
        # Remove X-Frame-Options entirely to allow embedding
        if 'X-Frame-Options' in response.headers:
            del response.headers['X-Frame-Options']

        # Allow embedding from SignPresenter and same origin (more secure than wildcard)
        # This allows: localhost, pythonanywhere.com, and signpresenter.com
        allowed_origins = "'self' https://*.signpresenter.com https://*.pythonanywhere.com http://localhost:* http://127.0.0.1:*"
        response.headers['Content-Security-Policy'] = f"frame-ancestors {allowed_origins};"
        response.headers['Access-Control-Allow-Origin'] = '*'

        return response

    except Exception as e:
        logger.error(f"Display page error: {str(e)}")
        flash('An error occurred while loading the display page.')
        return redirect(url_for('index'))

@app.route('/get_firefighter_logs/<fireman_number>')
def get_firefighter_logs(fireman_number):
    """API endpoint to get firefighter logs"""
    if not session.get('logged_in'):
        return jsonify({'error': 'Not authorized'}), 401

    firefighter = db_helpers.get_firefighter_by_number(fireman_number)

    if not firefighter:
        return jsonify({'error': 'Firefighter not found'}), 404

    logs = db_helpers.get_firefighter_logs(fireman_number)

    # Format logs for display
    formatted_logs = []
    for log in logs:
        time_in = datetime.fromisoformat(log['time_in'])
        if time_in.tzinfo is None:
            time_in = pytz.utc.localize(time_in)
        time_in = time_in.astimezone(central)

        hours = 'N/A'
        time_out_display = 'Still Active'

        if log['time_out']:
            time_out = datetime.fromisoformat(log['time_out'])
            if time_out.tzinfo is None:
                time_out = pytz.utc.localize(time_out)
            time_out = time_out.astimezone(central)
            time_out_display = time_out.strftime('%Y-%m-%d %H:%M:%S')

            hours_val = log.get('hours_worked') or log.get('manual_added_hours')
            if hours_val:
                hours = f"{hours_val:.2f}"

        formatted_logs.append({
            'type': log['type'],
            'time_in': time_in.strftime('%Y-%m-%d %H:%M:%S'),
            'time_out': time_out_display,
            'hours': hours
        })

    return jsonify({
        'logs': formatted_logs,
        'name': firefighter['full_name']
    })

# ========== VEHICLE INSPECTION ROUTES ==========

@app.route('/inspections')
def inspections_menu():
    """Vehicle inspections menu - select a vehicle"""
    # Get optional station filter from query parameters
    station_param = request.args.get('station')

    # Get all stations for the filter dropdown
    stations = db_helpers.get_all_stations()

    # Determine station_id for filtering
    if station_param is None:
        # First visit - no parameter, default to first station
        if stations:
            station_id = stations[0]['id']
        else:
            station_id = None
    elif station_param == 'all' or station_param == '':
        # User selected "All Stations" - show all vehicles
        station_id = None
    else:
        # User selected a specific station
        station_id = int(station_param)

    # Get vehicles (filtered by station if specified)
    vehicles = db_helpers.get_vehicles_needing_inspection(station_id=station_id)

    return render_template('inspections_menu.html',
                         vehicles=vehicles,
                         stations=stations,
                         selected_station=station_id)

@app.route('/inspect/<int:vehicle_id>')
def inspect_vehicle(vehicle_id):
    """Start inspection for a specific vehicle"""
    vehicle = db_helpers.get_vehicle_by_id(vehicle_id)

    if not vehicle:
        flash('Vehicle not found!')
        return redirect(url_for('inspections_menu'))

    # Get vehicle-specific checklist items
    checklist_items = db_helpers.get_vehicle_checklist(vehicle_id)
    firefighters = db_helpers.get_all_firefighters()

    return render_template('inspect_vehicle.html',
                         vehicle=vehicle,
                         checklist_items=checklist_items,
                         firefighters=firefighters)

@app.route('/submit_inspection', methods=['POST'])
def submit_inspection():
    """Submit a vehicle inspection"""
    try:
        vehicle_id = int(request.form['vehicle_id'])
        inspector_number = request.form.get('inspector_number', '')
        additional_notes = request.form.get('additional_notes', '')

        # Validate that inspector number is provided
        if not inspector_number:
            flash('Please select your firefighter number before submitting.')
            return redirect(url_for('inspect_vehicle', vehicle_id=vehicle_id))

        # Get inspector ID
        inspector = db_helpers.get_firefighter_by_number(inspector_number)

        if not inspector:
            flash('Invalid firefighter number selected.')
            return redirect(url_for('inspect_vehicle', vehicle_id=vehicle_id))

        inspector_id = inspector['id']

        # Collect inspection results
        inspection_results = []
        checklist_items = db_helpers.get_inspection_checklist()

        for item in checklist_items:
            item_id = item['id']
            status = request.form.get(f'item_{item_id}', 'pass')
            notes = request.form.get(f'notes_{item_id}', '')

            inspection_results.append({
                'item_id': item_id,
                'status': status,
                'notes': notes
            })

        # Save inspection
        success, result = db_helpers.create_vehicle_inspection(
            vehicle_id,
            inspector_id,
            inspection_results,
            additional_notes
        )

        if success:
            vehicle = db_helpers.get_vehicle_by_id(vehicle_id)
            flash(f'Inspection for {vehicle["name"]} completed successfully!')
            logger.info(f"Vehicle inspection completed: {vehicle['name']}")
            return redirect(url_for('inspections_menu'))
        else:
            flash(f'Error saving inspection: {result}')
            return redirect(url_for('inspect_vehicle', vehicle_id=vehicle_id))

    except Exception as e:
        logger.error(f"Submit inspection error: {str(e)}")
        flash('An error occurred while submitting the inspection.')
        return redirect(url_for('inspections_menu'))

@app.route('/inspection_history/<int:vehicle_id>')
def inspection_history(vehicle_id):
    """View inspection history for a vehicle"""
    vehicle = db_helpers.get_vehicle_by_id(vehicle_id)

    if not vehicle:
        flash('Vehicle not found!')
        return redirect(url_for('inspections_menu'))

    history = db_helpers.get_vehicle_inspection_history(vehicle_id)

    return render_template('inspection_history.html',
                         vehicle=vehicle,
                         history=history)

# ========== MAINTENANCE ROUTES ==========

@app.route('/maintenance')
def maintenance_menu():
    """Maintenance menu - select a vehicle"""
    vehicles = db_helpers.get_all_vehicles()
    alerts = db_helpers.get_all_alerts()

    # Create a set of vehicle IDs that have failed inspections
    failed_vehicle_ids = {v['id'] for v in alerts['inspections_failed']}

    # Add failed inspection info to each vehicle
    for vehicle in vehicles:
        vehicle['has_failed_inspection'] = vehicle['id'] in failed_vehicle_ids
        # Find the specific failure info if it exists
        for failed in alerts['inspections_failed']:
            if failed['id'] == vehicle['id']:
                vehicle['failed_inspection_notes'] = failed['notes']
                vehicle['failed_date'] = failed['failed_date']
                break

    return render_template('maintenance_menu.html', vehicles=vehicles)

@app.route('/maintenance/<int:vehicle_id>')
def maintenance_form(vehicle_id):
    """Maintenance work order form for a specific vehicle"""
    vehicle = db_helpers.get_vehicle_by_id(vehicle_id)

    if not vehicle:
        flash('Vehicle not found!')
        return redirect(url_for('maintenance_menu'))

    firefighters = db_helpers.get_all_firefighters()

    return render_template('maintenance_form.html',
                         vehicle=vehicle,
                         firefighters=firefighters)

@app.route('/submit_maintenance', methods=['POST'])
def submit_maintenance():
    """Handle maintenance work order submission"""
    try:
        vehicle_id = int(request.form['vehicle_id'])
        work_type = request.form['work_type']
        performed_by = request.form['performed_by']
        performed_date = request.form['performed_date']
        cost = request.form.get('cost', '')
        parts_used = request.form.get('parts_used', '')
        notes = request.form.get('notes', '')
        firefighter_number = request.form.get('firefighter_number', '')

        # Convert cost to float if provided
        cost_value = None
        if cost and cost.strip():
            try:
                cost_value = float(cost)
            except ValueError:
                flash('Invalid cost value')
                return redirect(url_for('maintenance_form', vehicle_id=vehicle_id))

        # Get firefighter ID if provided
        firefighter_id = None
        if firefighter_number:
            firefighter = db_helpers.get_firefighter_by_number(firefighter_number)
            if firefighter:
                firefighter_id = firefighter['id']

        # Create maintenance record
        success, result = db_helpers.create_maintenance_record(
            vehicle_id=vehicle_id,
            work_type=work_type,
            performed_by=performed_by,
            performed_date=performed_date,
            cost=cost_value,
            parts_used=parts_used,
            notes=notes,
            firefighter_id=firefighter_id
        )

        if success:
            vehicle = db_helpers.get_vehicle_by_id(vehicle_id)
            flash(f'Maintenance record for {vehicle["name"]} saved successfully!')
            logger.info(f"Maintenance record created for {vehicle['name']}: {work_type}")
            return redirect(url_for('maintenance_menu'))
        else:
            flash(f'Error saving maintenance record: {result}')
            return redirect(url_for('maintenance_form', vehicle_id=vehicle_id))

    except Exception as e:
        logger.error(f"Submit maintenance error: {str(e)}")
        flash('An error occurred while submitting the maintenance record.')
        return redirect(url_for('maintenance_menu'))

@app.route('/maintenance_history/<int:vehicle_id>')
def maintenance_history(vehicle_id):
    """View maintenance history for a vehicle"""
    vehicle = db_helpers.get_vehicle_by_id(vehicle_id)

    if not vehicle:
        flash('Vehicle not found!')
        return redirect(url_for('maintenance_menu'))

    history = db_helpers.get_maintenance_records_for_vehicle(vehicle_id)

    return render_template('maintenance_history.html',
                         vehicle=vehicle,
                         history=history)

# ========== INVENTORY ROUTES ==========

@app.route('/inventory')
def inventory_menu():
    """Inventory management menu - select station or vehicle"""
    stations = db_helpers.get_all_stations()
    vehicles = db_helpers.get_all_vehicles()

    return render_template('inventory_menu.html',
                         stations=stations,
                         vehicles=vehicles)

@app.route('/inventory/station/<int:station_id>')
def station_inventory(station_id):
    """View and manage inventory for a specific station"""
    station = db_helpers.get_station_by_id(station_id) if hasattr(db_helpers, 'get_station_by_id') else None

    # If get_station_by_id doesn't exist, get station info from the list
    if not station:
        stations = db_helpers.get_all_stations()
        station = next((s for s in stations if s['id'] == station_id), None)

    if not station:
        flash('Station not found!')
        return redirect(url_for('inventory_menu'))

    inventory = db_helpers.get_station_inventory(station_id)
    vehicles = db_helpers.get_vehicles_by_station(station_id)
    all_items = db_helpers.get_all_inventory_items()

    return render_template('station_inventory.html',
                         station=station,
                         inventory=inventory,
                         vehicles=vehicles,
                         all_items=all_items)

@app.route('/inventory/vehicle/<int:vehicle_id>')
def vehicle_inventory(vehicle_id):
    """View and manage inventory for a specific vehicle"""
    vehicle = db_helpers.get_vehicle_by_id(vehicle_id)

    if not vehicle:
        flash('Vehicle not found!')
        return redirect(url_for('inventory_menu'))

    inventory = db_helpers.get_vehicle_inventory(vehicle_id)
    all_items = db_helpers.get_all_inventory_items()

    return render_template('vehicle_inventory.html',
                         vehicle=vehicle,
                         inventory=inventory,
                         all_items=all_items)

@app.route('/inventory/add_to_station', methods=['POST'])
def add_to_station_inventory():
    """Add an item to station inventory"""
    try:
        station_id = int(request.form['station_id'])
        item_id = int(request.form['item_id'])
        quantity = int(request.form['quantity'])
        notes = request.form.get('notes', '')

        success, message = db_helpers.add_item_to_station(station_id, item_id, quantity, notes)

        if success:
            flash('Item added to station inventory successfully!')
        else:
            flash(f'Error adding item: {message}')

        return redirect(url_for('station_inventory', station_id=station_id))

    except Exception as e:
        logger.error(f"Add to station inventory error: {str(e)}")
        flash('An error occurred while adding the item.')
        return redirect(url_for('inventory_menu'))

@app.route('/inventory/add_to_vehicle', methods=['POST'])
def add_to_vehicle_inventory():
    """Add an item to vehicle inventory"""
    try:
        vehicle_id = int(request.form['vehicle_id'])
        item_id = int(request.form['item_id'])
        quantity = int(request.form['quantity'])
        notes = request.form.get('notes', '')

        success, message = db_helpers.add_item_to_vehicle(vehicle_id, item_id, quantity, notes)

        if success:
            flash('Item added to vehicle inventory successfully!')
        else:
            flash(f'Error adding item: {message}')

        return redirect(url_for('vehicle_inventory', vehicle_id=vehicle_id))

    except Exception as e:
        logger.error(f"Add to vehicle inventory error: {str(e)}")
        flash('An error occurred while adding the item.')
        return redirect(url_for('inventory_menu'))

@app.route('/inventory/update_station_quantity', methods=['POST'])
def update_station_quantity():
    """Update quantity of an item in station inventory"""
    try:
        station_inventory_id = int(request.form['station_inventory_id'])
        new_quantity = int(request.form['new_quantity'])
        station_id = int(request.form['station_id'])

        success, message = db_helpers.update_station_inventory_quantity(station_inventory_id, new_quantity)

        if success:
            flash('Quantity updated successfully!')
        else:
            flash(f'Error updating quantity: {message}')

        return redirect(url_for('station_inventory', station_id=station_id))

    except Exception as e:
        logger.error(f"Update station quantity error: {str(e)}")
        flash('An error occurred while updating the quantity.')
        return redirect(url_for('inventory_menu'))

@app.route('/inventory/update_vehicle_quantity', methods=['POST'])
def update_vehicle_quantity():
    """Update quantity of an item in vehicle inventory"""
    try:
        vehicle_inventory_id = int(request.form['vehicle_inventory_id'])
        new_quantity = int(request.form['new_quantity'])
        vehicle_id = int(request.form['vehicle_id'])

        success, message = db_helpers.update_vehicle_inventory_quantity(vehicle_inventory_id, new_quantity)

        if success:
            flash('Quantity updated successfully!')
        else:
            flash(f'Error updating quantity: {message}')

        return redirect(url_for('vehicle_inventory', vehicle_id=vehicle_id))

    except Exception as e:
        logger.error(f"Update vehicle quantity error: {str(e)}")
        flash('An error occurred while updating the quantity.')
        return redirect(url_for('inventory_menu'))

@app.route('/inventory/remove_from_station/<int:station_inventory_id>/<int:station_id>')
def remove_from_station(station_inventory_id, station_id):
    """Remove an item from station inventory"""
    success, message = db_helpers.remove_item_from_station(station_inventory_id)

    if success:
        flash('Item removed from station inventory!')
    else:
        flash(f'Error removing item: {message}')

    return redirect(url_for('station_inventory', station_id=station_id))

@app.route('/inventory/remove_from_vehicle/<int:vehicle_inventory_id>/<int:vehicle_id>')
def remove_from_vehicle(vehicle_inventory_id, vehicle_id):
    """Remove an item from vehicle inventory"""
    success, message = db_helpers.remove_item_from_vehicle(vehicle_inventory_id)

    if success:
        flash('Item removed from vehicle inventory!')
    else:
        flash(f'Error removing item: {message}')

    return redirect(url_for('vehicle_inventory', vehicle_id=vehicle_id))

@app.route('/inventory/create_item', methods=['POST'])
def create_inventory_item():
    """Create a new inventory item in the master catalog"""
    try:
        name = request.form['name']
        category = request.form['category']
        item_code = request.form.get('item_code', '')
        unit_of_measure = request.form.get('unit_of_measure', 'each')
        cost_per_unit = request.form.get('cost_per_unit', '')

        cost_value = None
        if cost_per_unit and cost_per_unit.strip():
            try:
                cost_value = float(cost_per_unit)
            except ValueError:
                flash('Invalid cost value')
                return redirect(request.referrer or url_for('inventory_menu'))

        success, result = db_helpers.create_inventory_item(
            name=name,
            category=category,
            item_code=item_code,
            unit_of_measure=unit_of_measure,
            cost_per_unit=cost_value
        )

        if success:
            flash(f'New item "{name}" created successfully!')
        else:
            flash(f'Error creating item: {result}')

        return redirect(request.referrer or url_for('inventory_menu'))

    except Exception as e:
        logger.error(f"Create inventory item error: {str(e)}")
        flash('An error occurred while creating the item.')
        return redirect(url_for('inventory_menu'))

# ========== VEHICLE MANAGEMENT ROUTES ==========

@app.route('/admin/vehicles')
def manage_vehicles():
    """Manage vehicles - view, add, edit"""
    vehicles = db_helpers.get_all_vehicles()
    stations = db_helpers.get_all_stations()
    return render_template('manage_vehicles.html', vehicles=vehicles, stations=stations)

@app.route('/admin/vehicle/create', methods=['POST'])
def create_vehicle():
    """Create a new vehicle"""
    try:
        # Vehicle code is now optional - will be auto-generated if empty
        vehicle_code = request.form.get('vehicle_code', '').strip()
        name = request.form['name']
        vehicle_type = request.form.get('vehicle_type', '')
        station_id = request.form.get('station_id', '')
        year = request.form.get('year', '')
        make = request.form.get('make', '')
        model = request.form.get('model', '')
        vin = request.form.get('vin', '')
        license_plate = request.form.get('license_plate', '')
        purchase_cost = request.form.get('purchase_cost', '')
        current_value = request.form.get('current_value', '')
        notes = request.form.get('notes', '')

        # Fluid specifications
        oil_type = request.form.get('oil_type', '')
        antifreeze_type = request.form.get('antifreeze_type', '')
        brake_fluid_type = request.form.get('brake_fluid_type', '')
        power_steering_fluid_type = request.form.get('power_steering_fluid_type', '')
        transmission_fluid_type = request.form.get('transmission_fluid_type', '')

        # Convert empty strings to None for optional fields
        station_id = int(station_id) if station_id else None
        year = int(year) if year else None
        purchase_cost = float(purchase_cost) if purchase_cost else None
        current_value = float(current_value) if current_value else None

        success, result = db_helpers.create_vehicle(
            vehicle_code=vehicle_code,
            name=name,
            vehicle_type=vehicle_type,
            station_id=station_id,
            year=year,
            make=make,
            model=model,
            vin=vin,
            license_plate=license_plate,
            purchase_cost=purchase_cost,
            current_value=current_value,
            notes=notes,
            oil_type=oil_type,
            antifreeze_type=antifreeze_type,
            brake_fluid_type=brake_fluid_type,
            power_steering_fluid_type=power_steering_fluid_type,
            transmission_fluid_type=transmission_fluid_type
        )

        if success:
            # Get the actual vehicle code that was used (auto-generated or provided)
            vehicle_id = result
            vehicle = db_helpers.get_vehicle_by_id(vehicle_id)
            actual_code = vehicle['code'] if vehicle else vehicle_code
            flash(f'Vehicle "{name}" created successfully with code {actual_code}!')
            logger.info(f"Vehicle created: {name} ({actual_code})")
        else:
            flash(f'Error creating vehicle: {result}')

        return redirect(url_for('manage_vehicles'))

    except Exception as e:
        logger.error(f"Create vehicle error: {str(e)}")
        flash('An error occurred while creating the vehicle.')
        return redirect(url_for('manage_vehicles'))

@app.route('/admin/vehicle/update/<int:vehicle_id>', methods=['POST'])
def update_vehicle(vehicle_id):
    """Update an existing vehicle"""
    try:
        vehicle_code = request.form.get('vehicle_code', '').strip()

        # If vehicle code is empty, keep the existing one
        if not vehicle_code:
            existing_vehicle = db_helpers.get_vehicle_by_id(vehicle_id)
            if existing_vehicle:
                vehicle_code = existing_vehicle['code']

        name = request.form['name']
        vehicle_type = request.form.get('vehicle_type', '')
        station_id = request.form.get('station_id', '')
        year = request.form.get('year', '')
        make = request.form.get('make', '')
        model = request.form.get('model', '')
        vin = request.form.get('vin', '')
        license_plate = request.form.get('license_plate', '')
        purchase_cost = request.form.get('purchase_cost', '')
        current_value = request.form.get('current_value', '')
        notes = request.form.get('notes', '')
        status = request.form.get('status', 'active')

        # Fluid specifications
        oil_type = request.form.get('oil_type', '')
        antifreeze_type = request.form.get('antifreeze_type', '')
        brake_fluid_type = request.form.get('brake_fluid_type', '')
        power_steering_fluid_type = request.form.get('power_steering_fluid_type', '')
        transmission_fluid_type = request.form.get('transmission_fluid_type', '')

        # Convert empty strings to None for optional fields
        station_id = int(station_id) if station_id else None
        year = int(year) if year else None
        purchase_cost = float(purchase_cost) if purchase_cost else None
        current_value = float(current_value) if current_value else None

        success, message = db_helpers.update_vehicle(
            vehicle_id=vehicle_id,
            vehicle_code=vehicle_code,
            name=name,
            vehicle_type=vehicle_type,
            station_id=station_id,
            year=year,
            make=make,
            model=model,
            vin=vin,
            license_plate=license_plate,
            purchase_cost=purchase_cost,
            current_value=current_value,
            notes=notes,
            status=status,
            oil_type=oil_type,
            antifreeze_type=antifreeze_type,
            brake_fluid_type=brake_fluid_type,
            power_steering_fluid_type=power_steering_fluid_type,
            transmission_fluid_type=transmission_fluid_type
        )

        if success:
            flash(f'Vehicle "{name}" updated successfully!')
            logger.info(f"Vehicle updated: {name} ({vehicle_code})")
            logger.info(f"Fluid types saved - Oil: {oil_type}, Antifreeze: {antifreeze_type}, Brake: {brake_fluid_type}")
        else:
            flash(f'Error updating vehicle: {message}')
            logger.error(f"Failed to update vehicle {vehicle_id}: {message}")

        return redirect(url_for('manage_vehicles'))

    except Exception as e:
        logger.error(f"Update vehicle error: {str(e)}")
        logger.error(f"Fluid data attempted - Oil: {oil_type}, Antifreeze: {antifreeze_type}, Brake: {brake_fluid_type}")
        flash(f'An error occurred while updating the vehicle: {str(e)}')
        return redirect(url_for('manage_vehicles'))

@app.route('/admin/vehicle/delete/<int:vehicle_id>', methods=['POST'])
def delete_vehicle(vehicle_id):
    """Delete a vehicle"""
    try:
        success, message = db_helpers.delete_vehicle(vehicle_id)

        if success:
            flash(message)
            logger.info(f"Vehicle deleted: ID {vehicle_id}")
        else:
            flash(f'Error deleting vehicle: {message}')

        return redirect(url_for('manage_vehicles'))

    except Exception as e:
        logger.error(f"Delete vehicle error: {str(e)}")
        flash('An error occurred while deleting the vehicle.')
        return redirect(url_for('manage_vehicles'))

@app.route('/admin/vehicle/<int:vehicle_id>/checklist')
def manage_vehicle_checklist(vehicle_id):
    """Manage checklist items for a specific vehicle"""
    vehicle = db_helpers.get_vehicle_by_id(vehicle_id)

    if not vehicle:
        flash('Vehicle not found!')
        return redirect(url_for('manage_vehicles'))

    # Get all available checklist items
    all_items = db_helpers.get_all_checklist_items()

    # Get items currently assigned to this vehicle
    assigned_items = db_helpers.get_vehicle_checklist(vehicle_id)
    assigned_item_ids = [item['id'] for item in assigned_items]

    return render_template('manage_vehicle_checklist.html',
                         vehicle=vehicle,
                         all_items=all_items,
                         assigned_items=assigned_items,
                         assigned_item_ids=assigned_item_ids)

@app.route('/admin/vehicle/<int:vehicle_id>/checklist/update', methods=['POST'])
def update_vehicle_checklist(vehicle_id):
    """Update checklist assignments for a vehicle"""
    try:
        vehicle = db_helpers.get_vehicle_by_id(vehicle_id)

        if not vehicle:
            flash('Vehicle not found!')
            return redirect(url_for('manage_vehicles'))

        # Get selected checklist item IDs from the form
        selected_item_ids = request.form.getlist('checklist_items')
        selected_item_ids = [int(item_id) for item_id in selected_item_ids]

        # Update the vehicle's checklist assignments
        success = db_helpers.assign_checklist_to_vehicle(vehicle_id, selected_item_ids)

        if success:
            flash(f'Checklist updated for {vehicle["name"]}! {len(selected_item_ids)} items assigned.')
            logger.info(f"Vehicle checklist updated: {vehicle['name']} - {len(selected_item_ids)} items")
        else:
            flash('Error updating checklist assignments.')

        return redirect(url_for('manage_vehicle_checklist', vehicle_id=vehicle_id))

    except Exception as e:
        logger.error(f"Update vehicle checklist error: {str(e)}")
        flash('An error occurred while updating the checklist.')
        return redirect(url_for('manage_vehicle_checklist', vehicle_id=vehicle_id))

# ========== STATION MANAGEMENT ROUTES ==========

@app.route('/admin/stations')
def manage_stations():
    """Manage stations - view and add"""
    stations = db_helpers.get_all_stations()
    return render_template('manage_stations.html', stations=stations)

@app.route('/admin/station/create', methods=['POST'])
def create_station():
    """Create a new fire station"""
    try:
        name = request.form['name']
        address = request.form.get('address', '')
        is_primary = request.form.get('is_primary') == 'on'
        notes = request.form.get('notes', '')

        success, result = db_helpers.create_station(
            name=name,
            address=address,
            is_primary=is_primary,
            notes=notes
        )

        if success:
            flash(f'Station "{name}" created successfully!')
            logger.info(f"Station created: {name}")
        else:
            flash(f'Error creating station: {result}')

        return redirect(url_for('manage_stations'))

    except Exception as e:
        logger.error(f"Create station error: {str(e)}")
        flash('An error occurred while creating the station.')
        return redirect(url_for('manage_stations'))

# ========== INSPECTION CHECKLIST MANAGEMENT ROUTES ==========

@app.route('/admin/checklist-items')
def manage_checklist_items():
    """Manage inspection checklist items"""
    items = db_helpers.get_all_checklist_items()
    return render_template('manage_checklist_items.html', items=items)

@app.route('/admin/checklist-item/create', methods=['POST'])
def create_checklist_item():
    """Create a new checklist item"""
    try:
        description = request.form['description']
        category = request.form.get('category', '')
        display_order = request.form.get('display_order', 0)

        success, result = db_helpers.create_checklist_item(
            description=description,
            category=category,
            display_order=int(display_order) if display_order else 0
        )

        if success:
            flash(f'Checklist item created successfully!')
            logger.info(f"Checklist item created: {description}")
        else:
            flash(f'Error creating checklist item: {result}')

        return redirect(url_for('manage_checklist_items'))

    except Exception as e:
        logger.error(f"Create checklist item error: {str(e)}")
        flash('An error occurred while creating the checklist item.')
        return redirect(url_for('manage_checklist_items'))

@app.route('/admin/checklist-item/toggle/<int:item_id>', methods=['POST'])
def toggle_checklist_item(item_id):
    """Toggle active status of a checklist item"""
    try:
        success = db_helpers.toggle_checklist_item(item_id)

        if success:
            flash('Checklist item status updated successfully!')
            logger.info(f"Checklist item {item_id} toggled")
        else:
            flash('Error updating checklist item status.')

        return redirect(url_for('manage_checklist_items'))

    except Exception as e:
        logger.error(f"Toggle checklist item error: {str(e)}")
        flash('An error occurred while updating the checklist item.')
        return redirect(url_for('manage_checklist_items'))

@app.route('/admin/checklist-item/delete/<int:item_id>', methods=['POST'])
def delete_checklist_item(item_id):
    """Delete a checklist item"""
    try:
        success = db_helpers.delete_checklist_item(item_id)

        if success:
            flash('Checklist item deleted successfully!')
            logger.info(f"Checklist item {item_id} deleted")
        else:
            flash('Error deleting checklist item.')

        return redirect(url_for('manage_checklist_items'))

    except Exception as e:
        logger.error(f"Delete checklist item error: {str(e)}")
        flash('An error occurred while deleting the checklist item.')
        return redirect(url_for('manage_checklist_items'))

@app.route('/admin/checklist-item/update-order/<int:item_id>', methods=['POST'])
def update_checklist_item_order(item_id):
    """Update the display order of a checklist item"""
    try:
        display_order = request.form.get('display_order', 0)
        success = db_helpers.update_checklist_item_order(item_id, int(display_order))

        if success:
            flash('Checklist item order updated successfully!')
            logger.info(f"Checklist item {item_id} order updated to {display_order}")
        else:
            flash('Error updating checklist item order.')

        return redirect(url_for('manage_checklist_items'))

    except Exception as e:
        logger.error(f"Update checklist item order error: {str(e)}")
        flash('An error occurred while updating the checklist item order.')
        return redirect(url_for('manage_checklist_items'))

# ========== DASHBOARD ROUTE ==========

@app.route('/dashboard')
def dashboard():
    """Main dashboard with statistics and charts"""
    try:
        logger.info("Loading dashboard stats...")
        stats = db_helpers.get_dashboard_stats()
        logger.info("Loading hours by day...")
        hours_by_day = db_helpers.get_hours_by_day(30)
        logger.info("Loading activity breakdown...")
        activity_breakdown = db_helpers.get_activity_breakdown()
        logger.info("Loading vehicle status...")
        vehicle_status = db_helpers.get_vehicle_status_summary()
        logger.info("Loading top performers...")
        top_performers = db_helpers.get_top_performers(10)
        logger.info("Loading alerts...")
        alerts = db_helpers.get_all_alerts()
        logger.info("Loading recent activity...")
        recent_activity = db_helpers.get_recent_activity(10)

        return render_template('dashboard.html',
                             stats=stats,
                             hours_by_day=hours_by_day,
                             activity_breakdown=activity_breakdown,
                             vehicle_status=vehicle_status,
                             top_performers=top_performers,
                             alerts=alerts,
                             recent_activity=recent_activity)

    except Exception as e:
        import traceback
        logger.error(f"Dashboard error: {str(e)}")
        logger.error(f"Traceback: {traceback.format_exc()}")
        flash(f'An error occurred while loading the dashboard: {str(e)}')
        return redirect(url_for('index'))

# ========== ALERTS DASHBOARD ROUTE ==========

@app.route('/alerts')
def alerts_dashboard():
    """Alerts dashboard showing all warnings and notifications"""
    alerts = db_helpers.get_all_alerts()
    display_settings = db_helpers.get_all_display_settings()
    stations = db_helpers.get_all_stations()
    return render_template('alerts_dashboard.html', alerts=alerts, display_settings=display_settings, stations=stations)

@app.route('/api/display-settings/toggle', methods=['POST'])
def toggle_display_setting():
    """API endpoint to toggle display settings (inventory/maintenance QR codes)"""
    try:
        data = request.get_json()
        setting_key = data.get('setting_key')
        new_value = data.get('value')

        if setting_key not in ['show_inventory_qr', 'show_maintenance_qr', 'show_inspections_qr']:
            return jsonify({'success': False, 'error': 'Invalid setting key'}), 400

        if new_value not in ['true', 'false']:
            return jsonify({'success': False, 'error': 'Invalid value'}), 400

        success = db_helpers.update_display_setting(setting_key, new_value)

        if success:
            return jsonify({'success': True, 'setting_key': setting_key, 'value': new_value})
        else:
            return jsonify({'success': False, 'error': 'Failed to update setting'}), 500

    except Exception as e:
        logger.error(f"Error toggling display setting: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500

# ========== BACKUP ROUTES ==========

@app.route('/admin/backups')
def backup_management():
    """Backup management page"""
    if not session.get('logged_in'):
        flash('Please log in first!')
        return redirect(url_for('admin'))

    try:
        # Get local backup status
        local_backup_status = db_helpers.get_backup_status()
        local_backups = db_helpers.list_database_backups()

        # Get Dropbox backup status
        dropbox_status = db_helpers.get_dropbox_backup_status()
        dropbox_backups_result = db_helpers.list_dropbox_backups()
        dropbox_backups = dropbox_backups_result.get('backups', [])

        return render_template('backup_management.html',
                             local_backup_status=local_backup_status,
                             local_backups=local_backups,
                             dropbox_status=dropbox_status,
                             dropbox_backups=dropbox_backups)
    except Exception as e:
        logger.error(f"Error loading backup management: {str(e)}")
        flash('An error occurred while loading backup information.')
        return redirect(url_for('admin_panel'))

@app.route('/admin/backups/create', methods=['POST'])
def create_backup():
    """Create a new database backup (and optionally upload to Dropbox)"""
    if not session.get('logged_in'):
        return jsonify({'success': False, 'error': 'Unauthorized'}), 401

    try:
        # Get upload_to_dropbox parameter from form
        upload_to_dropbox = request.form.get('upload_to_dropbox', 'false').lower() == 'true'

        result = db_helpers.create_database_backup()

        if result['success']:
            logger.info(f"Manual backup created: {result['backup_filename']}")

            # Optionally upload to Dropbox
            if upload_to_dropbox:
                dropbox_result = db_helpers.upload_backup_to_dropbox(result['backup_path'])
                if dropbox_result['success']:
                    result['dropbox_uploaded'] = True
                    result['dropbox_path'] = dropbox_result['dropbox_path']
                    logger.info(f"Backup uploaded to Dropbox: {dropbox_result['dropbox_path']}")
                    flash(f"Backup created and uploaded to Dropbox: {result['backup_filename']}")
                else:
                    result['dropbox_uploaded'] = False
                    result['dropbox_error'] = dropbox_result['error']
                    logger.warning(f"Backup created but Dropbox upload failed: {dropbox_result['error']}")
                    flash(f"Backup created locally but Dropbox upload failed: {result['backup_filename']}")
            else:
                flash(f"Backup created successfully: {result['backup_filename']}")

            return jsonify(result)
        else:
            logger.error(f"Backup creation failed: {result.get('error')}")
            flash(f"Backup failed: {result.get('error')}")
            return jsonify(result), 500

    except Exception as e:
        logger.error(f"Error creating backup: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/admin/backups/download/<filename>')
def download_backup(filename):
    """Download a backup file"""
    if not session.get('logged_in'):
        flash('Please log in first!')
        return redirect(url_for('admin'))

    try:
        # Security: validate filename to prevent directory traversal
        if '..' in filename or '/' in filename:
            flash('Invalid filename.')
            return redirect(url_for('backup_management'))

        backup_dir = os.path.join(os.path.dirname(db_helpers.DATABASE_PATH), 'backups')
        backup_path = os.path.join(backup_dir, filename)

        if not os.path.exists(backup_path):
            flash('Backup file not found.')
            return redirect(url_for('backup_management'))

        return send_file(backup_path, as_attachment=True, download_name=filename)

    except Exception as e:
        logger.error(f"Error downloading backup: {str(e)}")
        flash('An error occurred while downloading the backup.')
        return redirect(url_for('backup_management'))

@app.route('/admin/backups/cleanup', methods=['POST'])
def cleanup_backups():
    """Cleanup old backups, keeping only the most recent"""
    if not session.get('logged_in'):
        return jsonify({'success': False, 'error': 'Unauthorized'}), 401

    try:
        keep_count = int(request.form.get('keep_count', 10))
        result = db_helpers.cleanup_old_backups(keep_count)

        if result['success']:
            logger.info(f"Backup cleanup: {result['deleted_count']} backups deleted")
            flash(result['message'])
            return jsonify(result)
        else:
            flash(f"Cleanup failed: {result.get('error')}")
            return jsonify(result), 500

    except Exception as e:
        logger.error(f"Error during backup cleanup: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500

# ========== REPORTS ROUTES ==========

@app.route('/reports')
def reports_menu():
    """Reports menu page"""
    return render_template('reports_menu.html')

@app.route('/reports/hours', methods=['GET', 'POST'])
def hours_report():
    """Firefighter hours report"""
    try:
        start_date = request.args.get('start_date') or request.form.get('start_date')
        end_date = request.args.get('end_date') or request.form.get('end_date')
        firefighter_id = request.args.get('firefighter_id') or request.form.get('firefighter_id')
        export_format = request.args.get('export')

        if firefighter_id:
            firefighter_id = int(firefighter_id) if firefighter_id else None

        report_data = db_helpers.get_hours_report(start_date, end_date, firefighter_id)

        # Export to Excel
        if export_format == 'excel':
            wb = Workbook()
            ws = wb.active
            ws.title = "Hours Report"

            # Header styling
            header_fill = PatternFill(start_color="1F4788", end_color="1F4788", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)

            # Headers
            headers = ['Date', 'Firefighter #', 'Name', 'Activity', 'Time In', 'Time Out', 'Hours']
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')

            # Data rows
            for row_num, entry in enumerate(report_data['data'], 2):
                ws.cell(row=row_num, column=1, value=entry['date'])
                ws.cell(row=row_num, column=2, value=entry['firefighter_number'])
                ws.cell(row=row_num, column=3, value=entry['firefighter_name'])
                ws.cell(row=row_num, column=4, value=entry['activity'])
                ws.cell(row=row_num, column=5, value=entry['time_in'])
                ws.cell(row=row_num, column=6, value=entry['time_out'] or 'Active')
                ws.cell(row=row_num, column=7, value=entry['hours'])

            # Total row
            total_row = len(report_data['data']) + 2
            ws.cell(row=total_row, column=6, value='TOTAL:').font = Font(bold=True)
            ws.cell(row=total_row, column=7, value=report_data['total_hours']).font = Font(bold=True)

            # Adjust column widths
            for col_num in range(1, 8):
                ws.column_dimensions[get_column_letter(col_num)].width = 18

            # Save to BytesIO
            output = BytesIO()
            wb.save(output)
            output.seek(0)

            filename = f"hours_report_{datetime.now().strftime('%Y%m%d')}.xlsx"
            return send_file(output, download_name=filename, as_attachment=True,
                           mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

        # Get all firefighters for filter dropdown
        firefighters = db_helpers.get_all_firefighters()

        return render_template('report_hours.html',
                             report=report_data,
                             firefighters=firefighters,
                             start_date=start_date,
                             end_date=end_date,
                             selected_firefighter=firefighter_id)

    except Exception as e:
        logger.error(f"Hours report error: {str(e)}")
        flash('An error occurred while generating the report.')
        return redirect(url_for('reports_menu'))

@app.route('/reports/firefighter-summary', methods=['GET', 'POST'])
def firefighter_summary_report():
    """Firefighter summary report"""
    try:
        start_date = request.args.get('start_date') or request.form.get('start_date')
        end_date = request.args.get('end_date') or request.form.get('end_date')
        export_format = request.args.get('export')

        report_data = db_helpers.get_firefighter_summary_report(start_date, end_date)

        # Export to Excel
        if export_format == 'excel':
            wb = Workbook()
            ws = wb.active
            ws.title = "Firefighter Summary"

            header_fill = PatternFill(start_color="1F4788", end_color="1F4788", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)

            headers = ['Firefighter #', 'Name', 'Days Worked', 'Sessions', 'Total Hours']
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')

            for row_num, entry in enumerate(report_data['data'], 2):
                ws.cell(row=row_num, column=1, value=entry['firefighter_number'])
                ws.cell(row=row_num, column=2, value=entry['firefighter_name'])
                ws.cell(row=row_num, column=3, value=entry['days_worked'])
                ws.cell(row=row_num, column=4, value=entry['sessions'])
                ws.cell(row=row_num, column=5, value=entry['total_hours'])

            total_row = len(report_data['data']) + 2
            ws.cell(row=total_row, column=4, value='TOTAL:').font = Font(bold=True)
            ws.cell(row=total_row, column=5, value=report_data['total_hours']).font = Font(bold=True)

            for col_num in range(1, 6):
                ws.column_dimensions[get_column_letter(col_num)].width = 18

            output = BytesIO()
            wb.save(output)
            output.seek(0)

            filename = f"firefighter_summary_{datetime.now().strftime('%Y%m%d')}.xlsx"
            return send_file(output, download_name=filename, as_attachment=True,
                           mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

        return render_template('report_firefighter_summary.html',
                             report=report_data,
                             start_date=start_date,
                             end_date=end_date)

    except Exception as e:
        logger.error(f"Firefighter summary report error: {str(e)}")
        flash('An error occurred while generating the report.')
        return redirect(url_for('reports_menu'))

@app.route('/reports/activity', methods=['GET', 'POST'])
def activity_report():
    """Activity breakdown report"""
    try:
        start_date = request.args.get('start_date') or request.form.get('start_date')
        end_date = request.args.get('end_date') or request.form.get('end_date')
        export_format = request.args.get('export')

        report_data = db_helpers.get_activity_report(start_date, end_date)

        # Export to Excel
        if export_format == 'excel':
            wb = Workbook()
            ws = wb.active
            ws.title = "Activity Report"

            header_fill = PatternFill(start_color="1F4788", end_color="1F4788", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)

            headers = ['Activity Type', 'Sessions', 'Unique Firefighters', 'Total Hours']
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')

            for row_num, entry in enumerate(report_data['data'], 2):
                ws.cell(row=row_num, column=1, value=entry['activity'])
                ws.cell(row=row_num, column=2, value=entry['sessions'])
                ws.cell(row=row_num, column=3, value=entry['unique_firefighters'])
                ws.cell(row=row_num, column=4, value=entry['total_hours'])

            total_row = len(report_data['data']) + 2
            ws.cell(row=total_row, column=3, value='TOTAL:').font = Font(bold=True)
            ws.cell(row=total_row, column=4, value=report_data['total_hours']).font = Font(bold=True)

            for col_num in range(1, 5):
                ws.column_dimensions[get_column_letter(col_num)].width = 22

            output = BytesIO()
            wb.save(output)
            output.seek(0)

            filename = f"activity_report_{datetime.now().strftime('%Y%m%d')}.xlsx"
            return send_file(output, download_name=filename, as_attachment=True,
                           mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

        return render_template('report_activity.html',
                             report=report_data,
                             start_date=start_date,
                             end_date=end_date)

    except Exception as e:
        logger.error(f"Activity report error: {str(e)}")
        flash('An error occurred while generating the report.')
        return redirect(url_for('reports_menu'))

@app.route('/reports/maintenance-costs', methods=['GET', 'POST'])
def maintenance_costs_report():
    """Maintenance costs report"""
    try:
        start_date = request.args.get('start_date') or request.form.get('start_date')
        end_date = request.args.get('end_date') or request.form.get('end_date')
        vehicle_id = request.args.get('vehicle_id') or request.form.get('vehicle_id')
        export_format = request.args.get('export')

        if vehicle_id:
            vehicle_id = int(vehicle_id) if vehicle_id else None

        report_data = db_helpers.get_maintenance_cost_report(start_date, end_date, vehicle_id)

        # Export to Excel
        if export_format == 'excel':
            wb = Workbook()
            ws = wb.active
            ws.title = "Maintenance Costs"

            header_fill = PatternFill(start_color="1F4788", end_color="1F4788", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)

            headers = ['Date', 'Vehicle', 'WO #', 'Title', 'Status', 'Priority', 'Labor Cost', 'Parts Cost', 'Total Cost']
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')

            for row_num, entry in enumerate(report_data['data'], 2):
                ws.cell(row=row_num, column=1, value=entry['reported_date'])
                ws.cell(row=row_num, column=2, value=f"{entry['vehicle_name']} ({entry['vehicle_code']})")
                ws.cell(row=row_num, column=3, value=entry['work_order_id'])
                ws.cell(row=row_num, column=4, value=entry['title'])
                ws.cell(row=row_num, column=5, value=entry['status'])
                ws.cell(row=row_num, column=6, value=entry['priority'])
                ws.cell(row=row_num, column=7, value=f"${entry['labor_cost']:.2f}")
                ws.cell(row=row_num, column=8, value=f"${entry['parts_cost']:.2f}")
                ws.cell(row=row_num, column=9, value=f"${entry['total_cost']:.2f}")

            total_row = len(report_data['data']) + 2
            ws.cell(row=total_row, column=6, value='TOTALS:').font = Font(bold=True)
            ws.cell(row=total_row, column=7, value=f"${report_data['total_labor_cost']:.2f}").font = Font(bold=True)
            ws.cell(row=total_row, column=8, value=f"${report_data['total_parts_cost']:.2f}").font = Font(bold=True)
            ws.cell(row=total_row, column=9, value=f"${report_data['total_cost']:.2f}").font = Font(bold=True)

            for col_num in range(1, 10):
                ws.column_dimensions[get_column_letter(col_num)].width = 16

            output = BytesIO()
            wb.save(output)
            output.seek(0)

            filename = f"maintenance_costs_{datetime.now().strftime('%Y%m%d')}.xlsx"
            return send_file(output, download_name=filename, as_attachment=True,
                           mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

        vehicles = db_helpers.get_all_vehicles()

        return render_template('report_maintenance_costs.html',
                             report=report_data,
                             vehicles=vehicles,
                             start_date=start_date,
                             end_date=end_date,
                             selected_vehicle=vehicle_id)

    except Exception as e:
        logger.error(f"Maintenance costs report error: {str(e)}", exc_info=True)
        flash(f'An error occurred while generating the report: {str(e)}')
        return redirect(url_for('reports_menu'))

@app.route('/reports/inventory-value', methods=['GET'])
def inventory_value_report():
    """Inventory value report"""
    try:
        export_format = request.args.get('export')

        report_data = db_helpers.get_inventory_value_report()

        # Export to Excel
        if export_format == 'excel':
            wb = Workbook()

            # Station Inventory Sheet
            ws1 = wb.active
            ws1.title = "Station Inventory"

            header_fill = PatternFill(start_color="1F4788", end_color="1F4788", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)

            headers = ['Location', 'Item Name', 'Category', 'Quantity', 'Cost/Unit', 'Total Value']
            for col_num, header in enumerate(headers, 1):
                cell = ws1.cell(row=1, column=col_num, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')

            for row_num, entry in enumerate(report_data['station_inventory'], 2):
                ws1.cell(row=row_num, column=1, value=entry['location'])
                ws1.cell(row=row_num, column=2, value=entry['item_name'])
                ws1.cell(row=row_num, column=3, value=entry['category'])
                ws1.cell(row=row_num, column=4, value=entry['quantity'])
                ws1.cell(row=row_num, column=5, value=f"${entry['cost_per_unit']:.2f}")
                ws1.cell(row=row_num, column=6, value=f"${entry['total_value']:.2f}")

            total_row = len(report_data['station_inventory']) + 2
            ws1.cell(row=total_row, column=5, value='TOTAL:').font = Font(bold=True)
            ws1.cell(row=total_row, column=6, value=f"${report_data['station_total']:.2f}").font = Font(bold=True)

            for col_num in range(1, 7):
                ws1.column_dimensions[get_column_letter(col_num)].width = 18

            # Vehicle Inventory Sheet
            ws2 = wb.create_sheet(title="Vehicle Inventory")

            for col_num, header in enumerate(headers, 1):
                cell = ws2.cell(row=1, column=col_num, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')

            for row_num, entry in enumerate(report_data['vehicle_inventory'], 2):
                ws2.cell(row=row_num, column=1, value=entry['location'])
                ws2.cell(row=row_num, column=2, value=entry['item_name'])
                ws2.cell(row=row_num, column=3, value=entry['category'])
                ws2.cell(row=row_num, column=4, value=entry['quantity'])
                ws2.cell(row=row_num, column=5, value=f"${entry['cost_per_unit']:.2f}")
                ws2.cell(row=row_num, column=6, value=f"${entry['total_value']:.2f}")

            total_row = len(report_data['vehicle_inventory']) + 2
            ws2.cell(row=total_row, column=5, value='TOTAL:').font = Font(bold=True)
            ws2.cell(row=total_row, column=6, value=f"${report_data['vehicle_total']:.2f}").font = Font(bold=True)

            for col_num in range(1, 7):
                ws2.column_dimensions[get_column_letter(col_num)].width = 18

            output = BytesIO()
            wb.save(output)
            output.seek(0)

            filename = f"inventory_value_{datetime.now().strftime('%Y%m%d')}.xlsx"
            return send_file(output, download_name=filename, as_attachment=True,
                           mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

        return render_template('report_inventory_value.html', report=report_data)

    except Exception as e:
        logger.error(f"Inventory value report error: {str(e)}")
        flash('An error occurred while generating the report.')
        return redirect(url_for('reports_menu'))

if __name__ == '__main__':
    # Get debug mode from environment variable (defaults to False for production)
    debug_mode = os.getenv('FLASK_DEBUG', 'False').lower() == 'true'
    app.run(debug=debug_mode, host='0.0.0.0', port=5001)
